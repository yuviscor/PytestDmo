import pytest 

import openpyxl

class DcData:

    varia = [{"name":"yuvraj", "email":"seek@gmail.com", "gender":"Male"},
             {"name": "yuvrajV", "email": "seekvvv@gmail.com", "gender": "Female"}]

    va = []
    @staticmethod
    def GetDataM():

        listOfData=[]

        wb = openpyxl.load_workbook("C:\\Users\\Yuvraj verma\\Desktop\\SeleniumTesting\\Excel Reading\\readfile.xlsx")

        s = wb.active
        
        print(s.cell(row=1, column=1).value)

        totR = s.max_row

        totC = s.max_column

        for i in range(2, totR+1):
            dict = {}

            for j in range(2, totC+1):

                dict[s.cell(row=1, column=j).value] = s.cell(row=i, column=j).value

            DcData.va.append(dict)
        print(dict)
        print(dict)

        print(dict)

        print(DcData.va)
        return DcData.va






